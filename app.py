import streamlit as st
import pandas as pd
import requests
import time
from urllib.parse import urlparse
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# --- CONFIGURATION ---
API_KEY = "6deab6fa95f607482fd4c2fcadda451b3b479048558fdaa76cb62e2f132349c6"
TARGET_DOMAIN = "kollegeapply.com"

COMPETITORS = {
    "shiksha": "shiksha.com",
    "collegedunia": "collegedunia.com",
    "collegedekho": "collegedekho.com"
}

OFFICIAL_PATTERNS = [
    "wikipedia.org", "wikimedia.org", "britannica.com", ".europa.eu"
]

AMBIGUOUS_QUERIES = {
    "cat": "CAT exam",
    "gmat": "GMAT exam",
    "gre": "GRE exam"
}

# --- HELPER FUNCTIONS ---
def is_official_site(url):
    return any(p in url.lower() for p in OFFICIAL_PATTERNS)

def extract_domain(url):
    try:
        return urlparse(url).netloc.lower().replace("www.", "")
    except:
        return ""

def domain_in_url(url, domain):
    try:
        return extract_domain(url) == domain
    except:
        return False

def get_ranking(organic_results, target_domain):
    for idx, r in enumerate(organic_results, start=1):
        link = r.get("link", "")
        domain = extract_domain(link)
        if domain == target_domain:
            return idx, link
    return "Not in Top 100", ""

def process_keywords(df_kw):
    results = []
    for kw in df_kw["KW"]:
        query = AMBIGUOUS_QUERIES.get(kw.strip().lower(), kw)
        params = {
            "engine": "google",
            "q": query,
            "api_key": API_KEY,
            "num": 100,
            "hl": "en",
            "gl": "in",
            "google_domain": "google.co.in",
            "device": "desktop"
        }

        try:
            resp = requests.get("https://serpapi.com/search", params=params)
            resp.raise_for_status()
            data = resp.json()

            organic = data.get("organic_results", [])
            row = {"keyword": kw}
            kr, ku = get_ranking(organic, TARGET_DOMAIN)
            row["kollegeapply_rank"] = kr
            row["kollegeapply_url"] = ku

            filtered = [r for r in organic if not is_official_site(r.get("link", ""))]
            for i in range(1, 4):
                if len(filtered) >= i:
                    row[f"rank_{i}_name"] = filtered[i - 1].get("title", "")
                    row[f"rank_{i}_url"] = filtered[i - 1].get("link", "")
                else:
                    row[f"rank_{i}_name"] = ""
                    row[f"rank_{i}_url"] = ""

            for name, dom in COMPETITORS.items():
                rnk, url = get_ranking(organic, dom)
                row[f"{name}_rank"] = rnk
                row[f"{name}_url"] = url

            row["paa_exists"] = "No"
            row["paa_kollegeapply"] = "No"
            checked_links = []

            featured = data.get("answer_box", {}) or data.get("featured_snippet", {})
            fs_link = featured.get("link", "")
            if fs_link:
                checked_links.append(fs_link)
                if domain_in_url(fs_link, TARGET_DOMAIN):
                    row["paa_exists"] = "Yes"
                    row["paa_kollegeapply"] = "Yes"

            paa_results = data.get("related_questions", [])
            if paa_results:
                row["paa_exists"] = "Yes"
                for item in paa_results:
                    link = ""
                    if "source" in item and "link" in item["source"]:
                        link = item["source"]["link"]
                    elif "answer" in item and "source" in item["answer"] and "link" in item["answer"]["source"]:
                        link = item["answer"]["source"]["link"]
                    if link:
                        checked_links.append(link)
                        if domain_in_url(link, TARGET_DOMAIN):
                            row["paa_kollegeapply"] = "Yes"
                            break

            row["paa_links_checked"] = "; ".join(checked_links)
            results.append(row)

        except Exception as e:
            st.warning(f"❌ Error processing keyword '{kw}': {e}")

        time.sleep(2)

    return pd.DataFrame(results)

def compare_ranks(row):
    try:
        cur = int(row["kollegeapply_rank"])
        prev = int(row["kollegeapply_rank_last"])
        if cur < prev:
            return "blue"
        elif cur > prev:
            return "red"
        else:
            return ""
    except:
        return ""

def rank_diff(row):
    try:
        cur = int(row["kollegeapply_rank"])
        prev = int(row["kollegeapply_rank_last"])
        diff = prev - cur
        if diff > 0:
            return f"+{diff}"
        elif diff < 0:
            return f"{diff}"
        else:
            return "0"
    except:
        return "N/A"

# --- STREAMLIT UI ---
st.set_page_config(page_title="KollegeApply Rank Checker", layout="wide")
st.title("🔍 KollegeApply Keyword Rank Tracker")
st.markdown("Upload current keyword list and last week's report to compare rankings, highlight changes, and view rank difference.")

uploaded_kw = st.file_uploader("Upload Current Keyword Excel", type=["xlsx"])
uploaded_old = st.file_uploader("Upload Previous Report Excel", type=["xlsx"])

if uploaded_kw and uploaded_old:
    df_kw = pd.read_excel(uploaded_kw, usecols=["KW"])
    df_previous = pd.read_excel(uploaded_old)

    with st.spinner("🔄 Processing keywords using SerpAPI..."):
        df_current = process_keywords(df_kw)

        df_merged = df_current.merge(
            df_previous[["keyword", "kollegeapply_rank"]],
            on="keyword",
            how="left",
            suffixes=("", "_last")
        )

        df_merged["rank_change_diff"] = df_merged.apply(rank_diff, axis=1)
        df_merged["rank_change_color"] = df_merged.apply(compare_ranks, axis=1)

        # Reorder columns: rank_change_diff immediately after kollegeapply_rank
        cols = df_merged.columns.tolist()
        if "kollegeapply_rank" in cols:
            idx = cols.index("kollegeapply_rank")
            reordered = (
                cols[:idx + 1]
                + ["rank_change_diff"]
                + ["kollegeapply_rank_last", "rank_change_color"]
                + [col for col in cols if col not in ("kollegeapply_rank_last", "rank_change_diff", "rank_change_color") and col not in cols[:idx + 1]]
            )
            df_merged = df_merged[reordered]

        st.success("✅ Completed. Previewing first 10 rows:")
        st.dataframe(df_merged.head(10))

        # Export to Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Rank Comparison"
        for r in dataframe_to_rows(df_merged, index=False, header=True):
            ws.append(r)

        rank_col_index = list(df_merged.columns).index("kollegeapply_rank") + 1
        color_map = {
            "red": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
            "blue": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        }

        for i, row in enumerate(df_merged.itertuples(), start=2):
            color = getattr(row, "rank_change_color")
            if color in color_map:
                ws.cell(row=i, column=rank_col_index).fill = color_map[color]

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        file_name = f"rank_comparison_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button("📥 Download Excel", data=output, file_name=file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
